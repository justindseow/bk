from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BANK_CLOSING_BALANCE = 48320
BOOK_BALANCE_BEFORE_BANK_ONLY = 42595
CASH_ACCOUNT = {"code": "1020", "name": "CIMB Current Account"}

NAVY = "1F2D3D"
BLUE = "2E4057"
GREEN = "0F766E"
GREEN_FILL = "E6F7F1"
LIGHT_FILL = "F8F9FB"
GOLD_FILL = "FFF7E5"
WHITE = "FFFFFF"
BORDER = "DDE4EC"
CURRENCY_FORMAT = '"RM"#,##0.00;[Red]("RM"#,##0.00)'


def build_excel_workbook(session: dict[str, Any]) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)

    journal_lines = generate_journal_lines(session)
    validation = validate_session(session, journal_lines)
    reconciliation = calculate_reconciliation(session)

    write_journal_voucher(workbook.create_sheet("Journal Voucher"), session, journal_lines, validation)
    write_trial_balance(workbook.create_sheet("Trial Balance"), session, journal_lines)
    write_document_ledger(workbook.create_sheet("Document Ledger WP1"), session)
    write_bank_verification(workbook.create_sheet("Bank Verification WP2"), session, reconciliation)
    write_adjusting_entries(workbook.create_sheet("Adjusting Entries"), session)
    write_next_session_checklist(workbook.create_sheet("Next Session Checklist"), session)
    write_depreciation_schedule(workbook.create_sheet("Depreciation Schedule"), session)
    write_prepaid_schedule(workbook.create_sheet("Prepaid Schedule"), session)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def export_filename(session: dict[str, Any]) -> str:
    client = session.get("client", {})
    entity = safe_filename_part(client.get("entityName", "Client"))
    period = client.get("period", "Session")
    period_parts = period.split()
    if len(period_parts) == 2:
        period_part = f"{period_parts[0][:3]}_{period_parts[1]}"
    else:
        period_part = safe_filename_part(period)
    return f"MacroByte_BK_{entity}_{period_part}.xlsx"


def safe_filename_part(value: str) -> str:
    return "_".join(str(value).replace("/", " ").split())


def parse_account(account_text: str | None) -> dict[str, str]:
    if not account_text:
        return {"code": "", "name": ""}
    parts = str(account_text).split(" - ", 1)
    return {"code": parts[0].strip(), "name": parts[1].strip() if len(parts) > 1 else ""}


def cash_line(document: dict[str, Any], source: str, index: int) -> dict[str, Any]:
    is_cash_in = document.get("flow") == "IN"
    amount = number(document.get("amount"))
    return {
        "id": f"{document.get('id')}-cash-{index}",
        "documentId": document.get("id"),
        "date": document.get("date"),
        "description": f"{document.get('party')} - {'cash received' if is_cash_in else 'cash paid'}",
        "accountCode": CASH_ACCOUNT["code"],
        "accountName": CASH_ACCOUNT["name"],
        "debit": amount if is_cash_in else 0,
        "credit": 0 if is_cash_in else amount,
        "source": source,
    }


def posted_lines(document: dict[str, Any]) -> list[dict[str, Any]]:
    account = parse_account(document.get("glAccount"))
    if not account["code"]:
        return []
    amount = number(document.get("amount"))
    document_line = {
        "id": f"{document.get('id')}-doc-1",
        "documentId": document.get("id"),
        "date": document.get("date"),
        "description": document.get("party"),
        "accountCode": account["code"],
        "accountName": account["name"],
        "debit": amount if document.get("flow") == "OUT" else 0,
        "credit": amount if document.get("flow") == "IN" else 0,
        "source": "Doc",
    }
    return [cash_line(document, "Doc", 0), document_line] if document.get("flow") == "IN" else [
        document_line,
        cash_line(document, "Doc", 0),
    ]


def split_lines(session: dict[str, Any], document: dict[str, Any]) -> list[dict[str, Any]]:
    split = find_by(session.get("splitDecisions", []), "documentId", document.get("id"))
    if not split:
        return []

    lines = []
    for index, line in enumerate(split.get("lines", [])):
        amount = number(line.get("amount"))
        lines.append(
            {
                "id": f"{document.get('id')}-split-{line.get('id')}-{index}",
                "documentId": document.get("id"),
                "date": document.get("date"),
                "description": line.get("description"),
                "accountCode": line.get("accountCode"),
                "accountName": line.get("accountName"),
                "debit": amount if line.get("direction") == "DR" else 0,
                "credit": amount if line.get("direction") == "CR" else 0,
                "source": "Split",
            }
        )

    return [cash_line(document, "Split", 0), *lines] if document.get("flow") == "IN" else [
        *lines,
        cash_line(document, "Split", 0),
    ]


def reclassify_lines(session: dict[str, Any], document: dict[str, Any]) -> list[dict[str, Any]]:
    decision = find_by(session.get("reclassifyDecisions", []), "documentId", document.get("id"))
    if not decision:
        return []
    amount = number(document.get("amount"))
    if document.get("flow") == "OUT":
        debit, credit, cash_debit, cash_credit = amount, 0, 0, amount
    else:
        debit, credit, cash_debit, cash_credit = 0, amount, amount, 0

    description = (
        f"{document.get('party')} - capitalised"
        if decision.get("reclassifyType") == "Asset purchase"
        else f"{document.get('party')} - {decision.get('directorNature', '')}"
    )
    reclass_line = {
        "id": f"{document.get('id')}-reclass-1",
        "documentId": document.get("id"),
        "date": document.get("date"),
        "description": description,
        "accountCode": decision.get("accountCode"),
        "accountName": decision.get("accountName"),
        "debit": debit,
        "credit": credit,
        "source": "Reclassify",
    }
    bank_line = cash_line(document, "Reclassify", 0)
    bank_line["debit"] = cash_debit
    bank_line["credit"] = cash_credit
    return [bank_line, reclass_line] if document.get("flow") == "IN" else [reclass_line, bank_line]


def bank_plus_lines(session: dict[str, Any]) -> list[dict[str, Any]]:
    lines = []
    for entry in session.get("bankOnlyEntries", []):
        bank_row = find_by(session.get("bankRows", []), "id", entry.get("bankRowId"))
        if not bank_row:
            continue
        amount = number(bank_row.get("amount"))
        entry_line = {
            "id": f"{bank_row.get('id')}-bank-plus-entry",
            "documentId": bank_row.get("id"),
            "date": bank_row.get("date"),
            "description": entry.get("description"),
            "accountCode": entry.get("accountCode"),
            "accountName": entry.get("accountName"),
            "debit": amount if bank_row.get("direction") == "DR" else 0,
            "credit": amount if bank_row.get("direction") == "CR" else 0,
            "source": "Bank+",
        }
        cash = {
            "id": f"{bank_row.get('id')}-bank-plus-cash",
            "documentId": bank_row.get("id"),
            "date": bank_row.get("date"),
            "description": f"{bank_row.get('description')} - bank {'payment' if bank_row.get('direction') == 'DR' else 'receipt'}",
            "accountCode": CASH_ACCOUNT["code"],
            "accountName": CASH_ACCOUNT["name"],
            "debit": amount if bank_row.get("direction") == "CR" else 0,
            "credit": amount if bank_row.get("direction") == "DR" else 0,
            "source": "Bank+",
        }
        lines.extend([entry_line, cash] if bank_row.get("direction") == "DR" else [cash, entry_line])
    return lines


def adjusting_lines(session: dict[str, Any]) -> list[dict[str, Any]]:
    lines = []
    for entry in session.get("adjustingEntries", []):
        debit_account = parse_account(entry.get("debitAccount"))
        credit_account = parse_account(entry.get("creditAccount"))
        amount = number(entry.get("amount"))
        if not debit_account["code"] or not credit_account["code"] or amount <= 0:
            continue
        lines.extend(
            [
                {
                    "id": f"{entry.get('id')}-adj-dr",
                    "documentId": entry.get("id"),
                    "date": entry.get("date"),
                    "description": entry.get("description"),
                    "accountCode": debit_account["code"],
                    "accountName": debit_account["name"],
                    "debit": amount,
                    "credit": 0,
                    "source": "Adjusting",
                },
                {
                    "id": f"{entry.get('id')}-adj-cr",
                    "documentId": entry.get("id"),
                    "date": entry.get("date"),
                    "description": entry.get("description"),
                    "accountCode": credit_account["code"],
                    "accountName": credit_account["name"],
                    "debit": 0,
                    "credit": amount,
                    "source": "Adjusting",
                },
            ]
        )
    return lines


def generate_journal_lines(session: dict[str, Any]) -> list[dict[str, Any]]:
    lines: list[dict[str, Any]] = []
    for document in session.get("documents", []):
        if document.get("status") == "Split Done":
            lines.extend(split_lines(session, document))
        elif document.get("status") == "Reclassified":
            lines.extend(reclassify_lines(session, document))
        elif document.get("status") == "Posted":
            lines.extend(posted_lines(document))
    lines.extend(bank_plus_lines(session))
    lines.extend(adjusting_lines(session))
    return lines


def validate_session(session: dict[str, Any], journal_lines: list[dict[str, Any]]) -> dict[str, Any]:
    total_debit = round(sum(number(line.get("debit")) for line in journal_lines), 2)
    total_credit = round(sum(number(line.get("credit")) for line in journal_lines), 2)
    critical = 0
    if abs(total_debit - total_credit) > 0.01:
        critical += 1
    critical += sum(
        1
        for doc in session.get("documents", [])
        if doc.get("status") in {"Needs Split", "Reclassify", "Pending Review"} or not doc.get("glAccount")
    )
    critical += sum(
        1
        for row in session.get("bankRows", [])
        if row.get("status") in {"Needs Review", "Match Multiple"}
        or (row.get("status") == "New" and not find_by(session.get("bankOnlyEntries", []), "bankRowId", row.get("id")))
    )
    reconciliation = calculate_reconciliation(session)
    if abs(reconciliation["difference"]) > 0.01:
        critical += 1
    critical += sum(1 for accrual in session.get("priorAccruals", []) if accrual.get("status") == "Pending")
    critical += sum(1 for entry in session.get("adjustingEntries", []) if entry.get("status") == "Pending Review")
    return {
        "total_debit": total_debit,
        "total_credit": total_credit,
        "difference": round(total_debit - total_credit, 2),
        "critical": critical,
        "status": "Passed" if critical == 0 else "Review Required",
    }


def calculate_reconciliation(session: dict[str, Any]) -> dict[str, float]:
    outstanding = sum(
        number(item.get("amount"))
        for item in session.get("timingItems", [])
        if item.get("timingType") == "Outstanding cheque"
    )
    deposits = sum(
        number(item.get("amount"))
        for item in session.get("timingItems", [])
        if item.get("timingType") == "Deposit in transit"
    )
    bank_only = 0.0
    for entry in session.get("bankOnlyEntries", []):
        row = find_by(session.get("bankRows", []), "id", entry.get("bankRowId"))
        if row:
            bank_only += number(row.get("amount"))
    adjusted_bank = BANK_CLOSING_BALANCE - outstanding + deposits
    adjusted_book = BOOK_BALANCE_BEFORE_BANK_ONLY + bank_only
    return {
        "bank_closing": BANK_CLOSING_BALANCE,
        "outstanding": outstanding,
        "deposits": deposits,
        "adjusted_bank": adjusted_bank,
        "book_balance": BOOK_BALANCE_BEFORE_BANK_ONLY,
        "bank_only": bank_only,
        "adjusted_book": adjusted_book,
        "difference": round(adjusted_bank - adjusted_book, 2),
    }


def write_sheet_header(sheet, session: dict[str, Any], title: str, carry_forward: bool = False) -> None:
    client = session.get("client", {})
    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    sheet["A1"].font = Font(bold=True, color=WHITE, size=14)
    sheet["A1"].fill = PatternFill("solid", fgColor=GREEN if carry_forward else NAVY)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append(["Entity", client.get("entityName", ""), "Period", client.get("period", ""), "Prepared By", client.get("preparedBy", "")])
    sheet.append([])


def table_header(sheet, headers: list[str], carry_forward: bool = False) -> int:
    sheet.append(headers)
    row = sheet.max_row
    fill = PatternFill("solid", fgColor=GREEN if carry_forward else BLUE)
    for cell in sheet[row]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = f"A{row + 1}"
    return row


def write_journal_voucher(sheet, session: dict[str, Any], lines: list[dict[str, Any]], validation: dict[str, Any]) -> None:
    write_sheet_header(sheet, session, "Journal Voucher")
    client = session.get("client", {})
    sheet.append(["Voucher Reference", voucher_reference(client.get("period", "")), "Prepared Date", datetime.now().date()])
    sheet.append(["Validation Status", validation["status"], "Finalisation Status", "Finalised" if session.get("journalVoucherFinalised") else "Open"])
    sheet.append([])
    header_row = table_header(sheet, ["Date", "Source", "Reference", "Description", "Account Code", "Account Name", "Debit", "Credit", "Notes"])
    for line in lines:
        sheet.append(
            [
                line.get("date"),
                line.get("source"),
                reference_for_line(session, line),
                line.get("description"),
                line.get("accountCode"),
                line.get("accountName"),
                number(line.get("debit")) or None,
                number(line.get("credit")) or None,
                notes_for_line(session, line),
            ]
        )
    total_row = sheet.max_row + 1
    sheet.append(["", "", "", "", "", "Total", f"=SUM(G{header_row + 1}:G{total_row - 1})", f"=SUM(H{header_row + 1}:H{total_row - 1})", ""])
    sheet.append(["", "", "", "", "", "Difference", f"=G{total_row}-H{total_row}", "", ""])
    sheet.append([])
    sheet.append(["Sign-off", "", "", "", "", "", "", "", ""])
    sheet.append(["Prepared by", client.get("preparedBy", ""), "Reviewed by", "", "Approved by", "", "Date", "", "Notes"])
    style_table(sheet, header_row, sheet.max_row, 9)


def write_trial_balance(sheet, session: dict[str, Any], lines: list[dict[str, Any]]) -> None:
    write_sheet_header(sheet, session, "Trial Balance")
    balances: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})
    for line in lines:
        key = (line.get("accountCode", ""), line.get("accountName", ""))
        balances[key]["debit"] += number(line.get("debit"))
        balances[key]["credit"] += number(line.get("credit"))
    header_row = table_header(sheet, ["Account Code", "Account Name", "Debit", "Credit", "Net Movement"])
    for (code, name), amounts in sorted(balances.items()):
        debit = round(amounts["debit"], 2)
        credit = round(amounts["credit"], 2)
        sheet.append([code, name, debit or None, credit or None, round(debit - credit, 2)])
    total_row = sheet.max_row + 1
    sheet.append(["", "Total", f"=SUM(C{header_row + 1}:C{total_row - 1})", f"=SUM(D{header_row + 1}:D{total_row - 1})", f"=C{total_row}-D{total_row}"])
    sheet.append(["", "Difference", f"=C{total_row}-D{total_row}", "", ""])
    style_table(sheet, header_row, sheet.max_row, 5)


def write_document_ledger(sheet, session: dict[str, Any]) -> None:
    write_sheet_header(sheet, session, "Document Ledger WP1")
    header_row = table_header(sheet, ["Date", "Document Ref", "Vendor / Customer", "Document Type", "Amount", "GL Account", "Status", "Split / Reclassify Notes"])
    for doc in session.get("documents", []):
        split = find_by(session.get("splitDecisions", []), "documentId", doc.get("id"))
        decision = find_by(session.get("reclassifyDecisions", []), "documentId", doc.get("id"))
        notes = doc.get("note", "")
        if split:
            lines = "; ".join(
                f"{line.get('direction')} {line.get('accountCode')} {line.get('accountName')} RM{number(line.get('amount')):,.2f}"
                for line in split.get("lines", [])
            )
            notes = f"{notes} Split: {lines}".strip()
        if decision:
            notes = f"{notes} Reclassify: {decision.get('accountCode')} {decision.get('accountName')} {decision.get('note', '')}".strip()
        sheet.append([doc.get("date"), doc.get("docRef"), doc.get("party"), doc.get("docType"), number(doc.get("amount")), doc.get("glAccount"), doc.get("status"), notes])
    style_table(sheet, header_row, sheet.max_row, 8)


def write_bank_verification(sheet, session: dict[str, Any], reconciliation: dict[str, float]) -> None:
    write_sheet_header(sheet, session, "Bank Verification WP2")
    header_row = table_header(sheet, ["Date", "Bank Description", "Reference", "Money In", "Money Out", "Suggested Match", "Status", "Matched Document Ref", "Timing Item Note", "New Bank Entry GL"])
    for row in session.get("bankRows", []):
        match = find_by(session.get("bankMatches", []), "bankRowId", row.get("id"))
        timing = find_by(session.get("timingItems", []), "bankRowId", row.get("id"))
        bank_entry = find_by(session.get("bankOnlyEntries", []), "bankRowId", row.get("id"))
        matched_refs = ""
        if match:
            matched_refs = ", ".join(
                doc.get("docRef", "")
                for doc_id in match.get("documentIds", [])
                for doc in [find_by(session.get("documents", []), "id", doc_id)]
                if doc
            )
        suggested = ", ".join(row.get("suggestedDocumentIds", []) or [])
        sheet.append(
            [
                row.get("date"),
                row.get("description"),
                row.get("reference"),
                number(row.get("amount")) if row.get("direction") == "CR" else None,
                number(row.get("amount")) if row.get("direction") == "DR" else None,
                suggested,
                row.get("status"),
                matched_refs or row.get("matchedTo"),
                timing.get("note") if timing else "",
                f"{bank_entry.get('accountCode')} {bank_entry.get('accountName')}" if bank_entry else "",
            ]
        )
    style_table(sheet, header_row, sheet.max_row, 10)
    sheet.append([])
    sheet.append(["Reconciliation"])
    recon_start = sheet.max_row + 1
    for label, value in [
        ("Bank closing balance", reconciliation["bank_closing"]),
        ("Less: outstanding cheques", reconciliation["outstanding"]),
        ("Add: deposits in transit", reconciliation["deposits"]),
        ("Adjusted bank balance", reconciliation["adjusted_bank"]),
        ("Book balance before bank-only entries", reconciliation["book_balance"]),
        ("Add / less: new bank entries posted in WP2", reconciliation["bank_only"]),
        ("Adjusted book balance", reconciliation["adjusted_book"]),
        ("Difference", reconciliation["difference"]),
    ]:
        sheet.append([label, value])
    style_table(sheet, recon_start, sheet.max_row, 2)


def write_adjusting_entries(sheet, session: dict[str, Any]) -> None:
    write_sheet_header(sheet, session, "Adjusting Entries")
    header_row = table_header(sheet, ["Date", "Type", "Description", "Debit Account", "Credit Account", "Amount", "Reverse Next Month", "Status", "Notes"])
    for entry in session.get("adjustingEntries", []):
        sheet.append([entry.get("date"), entry.get("type"), entry.get("description"), entry.get("debitAccount"), entry.get("creditAccount"), number(entry.get("amount")), "Yes" if entry.get("reverseNextMonth") else "No", entry.get("status"), entry.get("notes", "")])
    for accrual in session.get("priorAccruals", []):
        sheet.append([accrual.get("reversalDate"), "Reversal", accrual.get("description"), accrual.get("creditAccount"), accrual.get("debitAccount"), number(accrual.get("originalAmount")), "No", accrual.get("status"), f"Original period: {accrual.get('originalPeriod')}"])
    style_table(sheet, header_row, sheet.max_row, 9)


def write_next_session_checklist(sheet, session: dict[str, Any]) -> None:
    write_sheet_header(sheet, session, "Next Session Checklist - Carry Forward", carry_forward=True)
    header_row = table_header(sheet, ["Category", "Priority", "Description", "Source Step", "Amount", "Due Timing", "Status", "Notes"], carry_forward=True)
    items = [*session.get("handoverItems", []), *session.get("manualHandoverItems", [])]
    if not items:
        sheet.append(["No handover items generated yet. Open Step 07 and reset the generated checklist.", "", "", "", "", "", "", ""])
    for handover_item in items:
        sheet.append([handover_item.get("category"), handover_item.get("priority"), handover_item.get("description"), handover_item.get("sourceStep"), number_or_blank(handover_item.get("amount")), handover_item.get("dueTiming"), handover_item.get("status"), "Manual item" if not handover_item.get("generated", True) else "Generated from session state"])
    style_table(sheet, header_row, sheet.max_row, 8, carry_forward=True)


def write_depreciation_schedule(sheet, session: dict[str, Any]) -> None:
    write_sheet_header(sheet, session, "Depreciation Schedule - Carry Forward", carry_forward=True)
    header_row = table_header(sheet, ["Asset Description", "Asset Account", "Purchase Date", "Cost", "Useful Life", "Monthly Depreciation", "Accumulated Depreciation", "Net Book Value", "Next Month Action"], carry_forward=True)
    items = session.get("depreciationSchedule", [])
    if not items:
        sheet.append(["No depreciation schedule items yet.", "", "", "", "", "", "", "", ""])
    for schedule in items:
        cost = number(schedule.get("cost"))
        monthly = number(schedule.get("monthlyDepreciation"))
        posted_months = 1 if schedule.get("status") == "Depreciation Posted" else 0
        net_book = cost - (monthly * posted_months)
        sheet.append([schedule.get("assetDescription"), schedule.get("assetAccount"), schedule.get("purchaseDate"), cost, schedule.get("usefulLifeMonths"), monthly, schedule.get("accumulatedDepreciationAccount"), net_book, "Post next month depreciation"])
    style_table(sheet, header_row, sheet.max_row, 9, carry_forward=True)


def write_prepaid_schedule(sheet, session: dict[str, Any]) -> None:
    write_sheet_header(sheet, session, "Prepaid Schedule - Carry Forward", carry_forward=True)
    header_row = table_header(sheet, ["Description", "Original Amount", "Expense This Month", "Prepaid Balance", "Start Month", "End Month", "Monthly Release", "Next Month Action"], carry_forward=True)
    prepayments = [split for split in session.get("splitDecisions", []) if split.get("splitType") == "Prepayment"]
    if not prepayments:
        sheet.append(["No prepaid items created from WP1 split entries.", "", "", "", "", "", "", ""])
    for split in prepayments:
        document = find_by(session.get("documents", []), "id", split.get("documentId")) or {}
        expense = sum(number(line.get("amount")) for line in split.get("lines", []) if "prepaid" not in line_text(line))
        prepaid = sum(number(line.get("amount")) for line in split.get("lines", []) if "prepaid" in line_text(line) or str(line.get("accountCode")) == "1120")
        monthly_release = round(prepaid / 11, 2) if prepaid else 0
        sheet.append([document.get("party") or document.get("docRef"), number(document.get("amount")), expense, prepaid, session.get("client", {}).get("period"), "To be reviewed", monthly_release, "Release next month prepaid portion"])
    style_table(sheet, header_row, sheet.max_row, 8, carry_forward=True)


def style_table(sheet, header_row: int, end_row: int, end_col: int, carry_forward: bool = False) -> None:
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet.iter_rows(min_row=header_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = CURRENCY_FORMAT
    if end_row > header_row:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(end_col)}{end_row}"
    for row_index in range(header_row + 1, end_row + 1):
        if row_index % 2 == 0:
            for cell in sheet[row_index][:end_col]:
                cell.fill = PatternFill("solid", fgColor=GREEN_FILL if carry_forward else LIGHT_FILL)
    for col in range(1, end_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = min(max_column_width(sheet, col), 42)


def max_column_width(sheet, col: int) -> int:
    width = 12
    for cell in sheet[get_column_letter(col)]:
        value = "" if cell.value is None else str(cell.value)
        width = max(width, min(len(value) + 2, 42))
    return width


def reference_for_line(session: dict[str, Any], line: dict[str, Any]) -> str:
    doc = find_by(session.get("documents", []), "id", line.get("documentId"))
    if doc:
        return doc.get("docRef", "")
    bank_row = find_by(session.get("bankRows", []), "id", line.get("documentId"))
    if bank_row:
        return bank_row.get("reference", "")
    return str(line.get("documentId", ""))


def notes_for_line(session: dict[str, Any], line: dict[str, Any]) -> str:
    if line.get("source") == "Bank+":
        return "Bank-only entry from WP2"
    if line.get("source") == "Adjusting":
        entry = find_by(session.get("adjustingEntries", []), "id", line.get("documentId"))
        return "Reverse next month" if entry and entry.get("reverseNextMonth") else "Period-end entry"
    if line.get("source") == "Split":
        return "Confirmed split line"
    if line.get("source") == "Reclassify":
        return "Confirmed reclassification"
    return "Posted from WP1"


def voucher_reference(period: str) -> str:
    return f"JV-{''.join(str(period).split()).upper()}-001"


def find_by(items: list[dict[str, Any]], key: str, value: Any) -> dict[str, Any] | None:
    return next((item for item in items if item.get(key) == value), None)


def line_text(line: dict[str, Any]) -> str:
    return f"{line.get('accountCode', '')} {line.get('accountName', '')} {line.get('description', '')}".lower()


def number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def number_or_blank(value: Any) -> float | str:
    if value is None or value == "":
        return ""
    return number(value)
